import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.uniprot_enzyme_explorer.models import EnzymeRecord
from src.uniprot_enzyme_explorer.ui_formatters import table_check_status


def export_to_csv(
    records: list[EnzymeRecord],
    output_file: Path,
    input_duplicate_ids: dict[str, int] | None = None,
) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    input_duplicate_ids = input_duplicate_ids or {}

    fieldnames = [
        "uniprot_id",
        "protein_name",
        "organism",
        "ec_number",
        "sequence_length",
        "molecular_weight_da",
        "reviewed_status",
        "status_identycznej_sekwencji",
        "grupa_identycznej_sekwencji",
        "czy_reprezentant",
        "hydrophobic_percent",
        "cysteine_count",
        "cysteine_percent",
        "most_common_amino_acid",
        "interpretation",
        "sequence",
        "sprawdzenie",
    ]
    with output_file.open(
        mode="w",
        encoding="utf-8-sig",
        newline="",
    ) as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
            delimiter=";",
        )

        writer.writeheader()

        for record in records:
            input_count = input_duplicate_ids.get(record.uniprot_id, 0)
            writer.writerow(
                {
                    "uniprot_id": record.uniprot_id,
                    "protein_name": record.protein_name,
                    "organism": record.organism,
                    "ec_number": record.ec_number,
                    "sequence_length": record.sequence_length,
                    "molecular_weight_da": record.molecular_weight,
                    "reviewed_status": record.reviewed_status,
                    "status_identycznej_sekwencji": record.qc_status,
                    "grupa_identycznej_sekwencji": record.duplicate_group or "",
                    "czy_reprezentant": (
                        "tak" if record.is_representative else "nie"
                    ),
                    "hydrophobic_percent": record.hydrophobic_percent,
                    "cysteine_count": record.cysteine_count,
                    "cysteine_percent": record.cysteine_percent,
                    "most_common_amino_acid": record.most_common_amino_acid,
                    "interpretation": record.interpretation,
                    "sequence": record.sequence,
                    "sprawdzenie": table_check_status(record, input_count),
                }
            )

    return output_file


def export_to_xlsx(
    records: list[EnzymeRecord],
    output_file: Path,
    input_duplicate_ids: dict[str, int] | None = None,
) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    input_duplicate_ids = input_duplicate_ids or {}

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Enzymy"

    headers = [
        "UniProt ID",
        "Nazwa białka",
        "Organizm",
        "Numer EC",
        "Długość sekwencji",
        "Masa cząsteczkowa",
        "Status rekordu",
        "Status identycznej sekwencji",
        "Grupa 100% identycznej sekwencji",
        "Czy reprezentant",
        "Aminokwasy hydrofobowe [%]",
        "Liczba cystein",
        "Cysteiny [%]",
        "Najczęstszy aminokwas",
        "Interpretacja",
        "Sekwencja",
        "Sprawdzenie",
    ]

    worksheet.append(headers)

    for record in records:
        input_count = input_duplicate_ids.get(record.uniprot_id, 0)
        worksheet.append(
            [
                record.uniprot_id,
                record.protein_name,
                record.organism,
                record.ec_number,
                record.sequence_length,
                record.molecular_weight,
                record.reviewed_status,
                record.qc_status,
                record.duplicate_group or "",
                "Tak" if record.is_representative else "Nie",
                record.hydrophobic_percent,
                record.cysteine_count,
                record.cysteine_percent,
                record.most_common_amino_acid,
                record.interpretation,
                record.sequence,
                table_check_status(record, input_count),
            ]
        )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    column_widths = {
        "A": 14,
        "B": 38,
        "C": 25,
        "D": 18,
        "E": 14,
        "F": 14,
        "G": 24,
        "H": 28,
        "I": 32,
        "J": 16,
        "K": 24,
        "L": 16,
        "M": 14,
        "N": 20,
        "O": 45,
        "P": 60,
        "Q": 28,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 30

        for cell in worksheet[row_number]:
            cell.alignment = Alignment(vertical="top")

        worksheet.cell(row_number, 7).alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    workbook.save(output_file)

    return output_file
